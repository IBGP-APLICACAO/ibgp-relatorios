"""
Gerador de Formulário de Aplicação de Prova
--------------------------------------------
Uso:
  python gerar_formulario_v3.py                          → PDF em branco (template)
  python gerar_formulario_v3.py planilha.xlsx            → ZIP com um PDF por escola
  python gerar_formulario_v3.py planilha.xlsx "Escola X" → PDF só daquela escola
"""

import sys, os, zipfile, re
from difflib import SequenceMatcher
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

W, H = A4

# ── Paleta IBGP ───────────────────────────────────────────────────
AZUL       = colors.HexColor("#393939")   # cinza escuro (header principal)
AZUL_MED   = colors.HexColor("#4A4A4A")   # cinza médio (subheaders manhã)
AZUL_CLARO = colors.HexColor("#F4F4F4")   # fundo dos campos
VERDE      = colors.HexColor("#CF3432")   # vermelho IBGP (turno tarde)
VERDE_MED  = colors.HexColor("#B02C2A")   # vermelho escuro
VERDE_CL   = colors.HexColor("#FDF4F4")   # fundo campos turno tarde
CINZA      = colors.HexColor("#555555")
CINZA_LN   = colors.HexColor("#CCCCCC")
BRANCO     = colors.white
VERMELHO   = colors.HexColor("#CF3432")   # acento vermelho
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
for _lp in [
    os.path.join(_SCRIPT_DIR, "IBGP.png"),
    os.path.join(_SCRIPT_DIR, "logo", "IBGP.png"),
    "/sessions/epic-determined-dijkstra/mnt/uploads/IBGP.png",
]:
    if os.path.exists(_lp):
        LOGO_PATH = _lp
        break
else:
    LOGO_PATH = os.path.join(_SCRIPT_DIR, "IBGP.png")

MARGIN     = 1.2*cm
INNER_W    = W - 2 * MARGIN
FH         = 0.52*cm   # field height padrão
GAP        = 0.1*cm    # gap entre campos numa linha
ROW_GAP    = 0.30*cm   # espaço entre linhas de campos


# ─────────────────────────────────────────────────────────────────
# Primitivas
# ─────────────────────────────────────────────────────────────────

def lbl(c, x, y, text, bold=False, size=7.5, col=CINZA):
    c.setFillColor(col)
    c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
    c.drawString(x, y, text)

def field_bg(c, x, y, w, h, color=AZUL_CLARO):
    c.setFillColor(color)
    c.setStrokeColor(CINZA_LN)
    c.setLineWidth(0.4)
    c.rect(x, y, w, h, fill=1, stroke=1)

def txt(c, x, y, w, h, name, multiline=False, prefill=""):
    """AcroForm text field. If prefill, draw static text instead."""
    if prefill:
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        # vertically center
        c.drawString(x + 0.15*cm, y + (h - 9/72*cm) / 2 + 0.05*cm, prefill)
    else:
        c.acroForm.textfield(
            name=name, x=x, y=y, width=w, height=h,
            borderStyle='underlined', borderWidth=0,
            fillColor=AZUL_CLARO, textColor=colors.black,
            fontSize=9,
            fieldFlags='multiline' if multiline else '',
            forceBorder=True,
        )

def txt_verde(c, x, y, w, h, name, multiline=False):
    c.acroForm.textfield(
        name=name, x=x, y=y, width=w, height=h,
        borderStyle='underlined', borderWidth=0,
        fillColor=VERDE_CL, textColor=colors.black,
        fontSize=9,
        fieldFlags='multiline' if multiline else '',
        forceBorder=True,
    )

def chk(c, x, y, name, size=10):
    c.acroForm.checkbox(
        name=name, x=x, y=y, size=size,
        borderStyle='solid', borderWidth=0.8,
        fillColor=AZUL_CLARO, forceBorder=True,
    )

def section_bar(c, y, title, color=AZUL, h=0.58*cm):
    c.setFillColor(color)
    c.rect(MARGIN, y - h, INNER_W, h, fill=1, stroke=0)
    c.setFillColor(BRANCO)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(MARGIN + 0.3*cm, y - h + 0.17*cm, title.upper())
    return y - h  # returns bottom of bar

def static_field(c, x, y, w, h, value, bg=AZUL_CLARO):
    """Pre-filled (locked) field shown as coloured box with text."""
    field_bg(c, x, y, w, h, color=bg)
    if value:
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        # clip text to fit
        max_chars = int(w / (0.14*cm))
        display = value[:max_chars]
        c.drawString(x + 0.15*cm, y + h/2 - 0.14*cm, display)

def page_footer(c, page_num, total=3):
    # Faixa vermelha fina
    c.setFillColor(VERMELHO)
    c.rect(0, 0.6*cm, W, 0.12*cm, fill=1, stroke=0)
    # Fundo cinza escuro
    c.setFillColor(AZUL)
    c.rect(0, 0, W, 0.6*cm, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica", 6.5)
    c.drawCentredString(W/2, 0.2*cm,
        "Formulário de Aplicação de Prova  •  IBGP — Instituto Brasileiro de Gestão e Pesquisa  •  Uso interno")
    c.drawRightString(W - MARGIN, 0.2*cm, f"Pág. {page_num}/{total}")


# ─────────────────────────────────────────────────────────────────
# Componentes compostos
# ─────────────────────────────────────────────────────────────────

def row_fields(c, y, items, prefills=None):
    """
    items: [(label, name, width_cm), ...]
    prefills: dict name→value (optional)
    Returns y after the row.
    """
    if prefills is None:
        prefills = {}
    x = MARGIN
    lbl_y = y
    field_y = y - FH
    for (label_text, name, w_cm) in items:
        wpt = w_cm * cm
        lbl(c, x, lbl_y, label_text)
        field_bg(c, x, field_y, wpt, FH)
        val = prefills.get(name, "")
        if val:
            static_field(c, x, field_y, wpt, FH, val)
        else:
            txt(c, x, field_y, wpt, FH, name)
        x += wpt + GAP
    return field_y - ROW_GAP


def witness_table(c, y, prefix, title):
    """3-row witness table. Returns bottom y."""
    col_n = INNER_W * 0.62
    col_c = INNER_W * 0.38
    row_h = 0.5*cm
    hdr_h = 0.38*cm

    lbl(c, MARGIN, y, title, bold=True, size=7.5, col=AZUL_MED)
    y -= 0.22*cm

    # header
    c.setFillColor(AZUL_MED)
    c.rect(MARGIN, y - hdr_h, INNER_W, hdr_h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7)
    c.drawString(MARGIN + 0.15*cm, y - hdr_h + 0.1*cm, "Nome completo da testemunha")
    c.drawString(MARGIN + col_n + 0.15*cm, y - hdr_h + 0.1*cm, "CPF")
    y -= hdr_h

    for i in range(3):
        field_bg(c, MARGIN,          y - row_h, col_n, row_h)
        field_bg(c, MARGIN + col_n,  y - row_h, col_c, row_h)
        txt(c, MARGIN,          y - row_h, col_n, row_h, f"{prefix}_nome_{i+1}")
        txt(c, MARGIN + col_n,  y - row_h, col_c, row_h, f"{prefix}_cpf_{i+1}")
        c.setStrokeColor(CINZA_LN); c.setLineWidth(0.3)
        c.line(MARGIN,          y - row_h, MARGIN + INNER_W, y - row_h)
        c.line(MARGIN + col_n,  y,         MARGIN + col_n,   y - row_h)
        y -= row_h

    c.setStrokeColor(CINZA_LN); c.setLineWidth(0.5)
    c.rect(MARGIN, y, INNER_W, row_h * 3 + hdr_h, fill=0, stroke=1)
    return y - 0.22*cm


def cargo_table(c, y, prefix, n_rows=7, cargo_data=None):
    """
    cargo_data: list of (sala, cargo_nome, previstos).
    6 colunas: Sala | Cargo/Função | Previstos | Presentes | Ausentes | Inclusão
    Sala é mesclada visualmente quando consecutiva (span).
    """
    # 6 colunas
    cols_w  = [INNER_W*0.17, INNER_W*0.36, INNER_W*0.12, INNER_W*0.12, INNER_W*0.115, INNER_W*0.115]
    headers = ["Sala", "Cargo / Função", "Previstos", "Presentes", "Ausentes", "Inclusão"]
    row_h   = 0.40*cm
    hdr_h   = 0.34*cm

    if cargo_data:
        n_rows = len(cargo_data)

    # Pré-calcular grupos de sala (para mescla visual)
    sala_spans = []   # (sala, start_idx, count)
    if cargo_data:
        i = 0
        while i < len(cargo_data):
            sala = cargo_data[i][0]
            j = i + 1
            while j < len(cargo_data) and cargo_data[j][0] == sala:
                j += 1
            sala_spans.append((sala, i, j - i))
            i = j
    sala_first_rows = {start: (sala, count) for sala, start, count in sala_spans}
    sala_skip_rows  = {k for _, start, count in sala_spans for k in range(start + 1, start + count)}

    # Header
    c.setFillColor(AZUL_MED)
    c.rect(MARGIN, y - hdr_h, INNER_W, hdr_h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7)
    cx = MARGIN
    for h_txt, cw in zip(headers, cols_w):
        c.drawString(cx + 0.1*cm, y - hdr_h + 0.08*cm, h_txt)
        cx += cw
    y -= hdr_h
    top_y = y   # referência para cálculo das posições das salas

    for i in range(n_rows):
        cx   = MARGIN
        bg   = AZUL_CLARO if i % 2 == 0 else colors.HexColor("#EBEBEB")
        nome = cargo_data[i][1] if cargo_data else ""
        prev = str(cargo_data[i][2]) if cargo_data else ""

        for j, cw in enumerate(cols_w):
            if cargo_data and j == 0:
                # Sala: sem fundo individual (será desenhada como span depois)
                cx += cw
                continue

            field_bg(c, cx, y - row_h, cw, row_h, color=bg)

            if cargo_data and j == 1:
                c.setFillColor(colors.black)
                fs = 7.0
                while fs > 5 and c.stringWidth(nome, "Helvetica", fs) > cw - 0.15*cm:
                    fs -= 0.3
                c.setFont("Helvetica", fs)
                c.drawString(cx + 0.08*cm, y - row_h + (row_h - fs/72*cm)/2 + 0.03*cm, nome)

            elif cargo_data and j == 2:
                c.setFillColor(colors.black); c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(cx + cw/2, y - row_h + (row_h - 8/72*cm)/2 + 0.03*cm, prev)

            else:
                # Presentes / Ausentes / Inclusão — editáveis
                txt(c, cx, y - row_h, cw, row_h, f"{prefix}_cargo{i+1}_col{j}")

            c.setStrokeColor(CINZA_LN); c.setLineWidth(0.25)
            if j < len(cols_w) - 1:
                c.line(cx + cw, y, cx + cw, y - row_h)
            cx += cw

        c.setStrokeColor(CINZA_LN); c.setLineWidth(0.25)
        c.line(MARGIN, y - row_h, MARGIN + INNER_W, y - row_h)
        y -= row_h

    # Desenhar colunas de sala mescladas (sobrepõe os fundos de linha)
    if cargo_data:
        sala_col_w = cols_w[0]
        for sala, start, count in sala_spans:
            span_h   = row_h * count
            span_top = top_y - start * row_h
            span_bot = span_top - span_h
            # Fundo cinza suave para destacar a sala
            c.setFillColor(colors.HexColor("#DCDCDC"))
            c.setStrokeColor(CINZA_LN); c.setLineWidth(0.4)
            c.rect(MARGIN, span_bot, sala_col_w, span_h, fill=1, stroke=1)
            # Texto da sala centralizado verticalmente
            c.setFillColor(colors.HexColor("#222222"))
            c.setFont("Helvetica-Bold", 6.5)
            text_y = span_bot + span_h/2 - 0.09*cm
            c.drawString(MARGIN + 0.08*cm, text_y, sala)
    else:
        # Tabela em branco: sala é campo editável normal
        for i in range(n_rows):
            bg = AZUL_CLARO if i % 2 == 0 else colors.HexColor("#EBEBEB")
            row_top = top_y - i * row_h
            field_bg(c, MARGIN, row_top - row_h, cols_w[0], row_h, color=bg)
            txt(c, MARGIN, row_top - row_h, cols_w[0], row_h, f"{prefix}_cargo{i+1}_col0")

    c.setStrokeColor(CINZA_LN); c.setLineWidth(0.5)
    c.rect(MARGIN, y, INNER_W, row_h * n_rows + hdr_h, fill=0, stroke=1)
    return y - 0.28*cm


def occurrence_item(c, y, label_text, name, h_desc=1.0*cm, n_candidates=0):
    """Sim/Não + N linhas de candidato + descrição."""
    lbl(c, MARGIN, y, label_text, bold=True, size=8, col=AZUL_MED)
    chk_x = MARGIN + 6.0*cm
    chk(c, chk_x, y - 0.32*cm, name + "_sim")
    lbl(c, chk_x + 0.38*cm, y, "Sim")
    chk(c, chk_x + 1.5*cm, y - 0.32*cm, name + "_nao")
    lbl(c, chk_x + 1.88*cm, y, "Não")
    y -= 0.42*cm

    if n_candidates > 0:
        w_sala = 2.8*cm
        w_insc = 3.5*cm
        w_nome = INNER_W - w_sala - w_insc - 0.2*cm
        # cabeçalho da mini-tabela de candidatos
        c.setFillColor(AZUL_MED)
        c.rect(MARGIN, y - 0.34*cm, INNER_W, 0.34*cm, fill=1, stroke=0)
        c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7)
        c.drawString(MARGIN + 0.1*cm,                    y - 0.26*cm, "Sala")
        c.drawString(MARGIN + w_sala + 0.1*cm,           y - 0.26*cm, "Nº de Inscrição")
        c.drawString(MARGIN + w_sala + w_insc + 0.2*cm,  y - 0.26*cm, "Nome do Candidato")
        y -= 0.34*cm
        for i in range(n_candidates):
            bg = AZUL_CLARO if i % 2 == 0 else colors.HexColor("#EBEBEB")
            x0 = MARGIN
            x1 = MARGIN + w_sala + 0.1*cm
            x2 = MARGIN + w_sala + w_insc + 0.1*cm
            field_bg(c, x0, y - FH, w_sala, FH, color=bg)
            field_bg(c, x1, y - FH, w_insc, FH, color=bg)
            field_bg(c, x2, y - FH, w_nome, FH, color=bg)
            txt(c, x0, y - FH, w_sala, FH, f"{name}_sala_{i+1}")
            txt(c, x1, y - FH, w_insc, FH, f"{name}_insc_{i+1}")
            txt(c, x2, y - FH, w_nome, FH, f"{name}_cand_{i+1}")
            c.setStrokeColor(CINZA_LN); c.setLineWidth(0.3)
            c.line(MARGIN, y - FH, MARGIN + INNER_W, y - FH)
            c.line(x1 - 0.05*cm, y, x1 - 0.05*cm, y - FH)
            c.line(x2 - 0.05*cm, y, x2 - 0.05*cm, y - FH)
            y -= FH
        c.setStrokeColor(CINZA_LN); c.setLineWidth(0.5)
        c.rect(MARGIN, y, INNER_W, FH * n_candidates + 0.34*cm, fill=0, stroke=1)
        y -= 0.2*cm

    lbl(c, MARGIN, y, "Descrição:", col=CINZA)
    y -= 0.2*cm
    field_bg(c, MARGIN, y - h_desc, INNER_W, h_desc)
    txt(c, MARGIN, y - h_desc, INNER_W, h_desc, name + "_desc", multiline=True)
    return y - h_desc - 0.32*cm


# ─────────────────────────────────────────────────────────────────
# MAIN GENERATOR
# ─────────────────────────────────────────────────────────────────

def generate_pdf(output_path, data=None, cargo_data=None):
    """
    data: dict with optional prefill values:
      concurso, edital, data_aplicacao, uf, municipio, local_escola,
      coord_ped, coord_log, coord_local, turno (MANHÃ / TARDE / AMBOS)
    """
    if data is None:
        data = {}

    def pf(key):
        return data.get(key, "")

    turno_flag = data.get("turno", "AMBOS").upper()
    has_manha = turno_flag in ("MANHÃ", "MANHA", "AMBOS", "")
    has_tarde = turno_flag in ("TARDE", "AMBOS", "")

    c = canvas.Canvas(output_path, pagesize=A4)
    c.setTitle("Formulário de Aplicação de Prova")

    # ═══════════════════════════════════════════════════════════════
    # PÁGINA 1 — Identificação + Turno Manhã
    # ═══════════════════════════════════════════════════════════════

    def draw_main_header(c, subtitle=None):
        """Header com logo IBGP + título + linha vermelha. Igual em todas as páginas."""
        hdr_h = 2.8*cm
        # Faixa vermelha no topo
        c.setFillColor(VERMELHO)
        c.rect(0, H - 0.28*cm, W, 0.28*cm, fill=1, stroke=0)
        # Logo — margem ajustada para não sobrepor a faixa vermelha
        try:
            from reportlab.lib.utils import ImageReader
            logo = ImageReader(LOGO_PATH)
            logo_w = 5.2*cm          # um pouco menor para o "P" ficar longe do divisor
            logo_h = logo_w * (711/2200)
            logo_y = H - hdr_h + (hdr_h - 0.28*cm - logo_h) / 2
            c.drawImage(logo, MARGIN, logo_y,
                        width=logo_w, height=logo_h, mask='auto')
        except Exception:
            c.setFillColor(AZUL); c.setFont("Helvetica-Bold", 14)
            c.drawString(MARGIN, H - 1.6*cm, "IBGP")
        # Divisor vertical vermelho — posicionado após a logo com folga
        div_x = MARGIN + 5.2*cm + 0.6*cm   # logo_w + 0.6 cm de espaço
        c.setStrokeColor(VERMELHO); c.setLineWidth(1.2)
        c.line(div_x, H - 0.45*cm, div_x, H - hdr_h + 0.3*cm)
        # Título à direita do divisor
        txt_x = div_x + 0.45*cm
        c.setFillColor(AZUL); c.setFont("Helvetica-Bold", 12)
        c.drawString(txt_x, H - 1.2*cm, "FORMULÁRIO DE APLICAÇÃO DE PROVA")
        if subtitle:
            c.setFillColor(VERMELHO); c.setFont("Helvetica-Bold", 8)
            c.drawString(txt_x, H - 1.7*cm, subtitle)
            c.setFillColor(CINZA); c.setFont("Helvetica", 7.5)
            c.drawString(txt_x, H - 2.1*cm,
                "Preencher no dia da aplicação ou em até 3 dias após  ·  Uso interno")
        else:
            c.setFillColor(CINZA); c.setFont("Helvetica", 7.5)
            c.drawString(txt_x, H - 1.75*cm,
                "Preencher no dia da aplicação ou em até 3 dias após")
            c.drawString(txt_x, H - 2.15*cm,
                "Entregar ao responsável pela consolidação")
        # Linha vermelha separadora inferior
        c.setStrokeColor(VERMELHO); c.setLineWidth(1.5)
        c.line(0, H - hdr_h, W, H - hdr_h)
        return H - hdr_h - 0.25*cm

    y = draw_main_header(c)

    # ── Identificação ──────────────────────────────────────────────
    y = section_bar(c, y, "1. Identificação")
    y -= 0.45*cm

    y = row_fields(c, y, [
        ("Concurso / Edital", "concurso", 11.5),
        ("Data da Aplicação", "data_aplicacao", 3.8),
        ("UF", "uf", 1.2),
    ], prefills={
        "concurso": pf("concurso") + (" — " + pf("edital") if pf("edital") else ""),
        "data_aplicacao": pf("data_aplicacao"),
        "uf": pf("uf"),
    })

    y = row_fields(c, y, [
        ("Município",       "municipio",    7.0),
        ("Local / Escola",  "local_escola", 9.7),
    ], prefills={
        "municipio":   pf("municipio"),
        "local_escola": pf("local_escola"),
    })

    y = row_fields(c, y, [
        ("Coordenador Pedagógico", "coord_ped",   5.6),
        ("Coordenador Logístico",  "coord_log",   5.6),
        ("Coordenação Local",      "coord_local", 5.5),
    ], prefills={
        "coord_ped":   pf("coord_ped"),
        "coord_log":   pf("coord_log"),
        "coord_local": pf("coord_local"),
    })

    # ── Apoio ──────────────────────────────────────────────────────
    lbl(c, MARGIN, y, "Houve apoio?")
    chk(c, MARGIN + 2.8*cm, y - 0.32*cm, "apoio_sim")
    lbl(c, MARGIN + 3.15*cm, y, "Sim")
    chk(c, MARGIN + 4.1*cm, y - 0.32*cm, "apoio_nao")
    lbl(c, MARGIN + 4.45*cm, y, "Não")
    lbl(c, MARGIN + 5.5*cm, y, "Nome do apoio:")
    field_bg(c, MARGIN + 7.6*cm, y - FH, 8.5*cm, FH)
    txt(c, MARGIN + 7.6*cm, y - FH, 8.5*cm, FH, "apoio_nome")
    y -= FH + ROW_GAP

    # ── Turno Manhã — Horários ─────────────────────────────────────
    y = section_bar(c, y, "2. Turno Manhã — Horários")
    y -= 0.45*cm

    y = row_fields(c, y, [
        ("Chegada da equipe",        "m_chegada",   4.2),
        ("Abertura do portão",       "m_ab_portao", 4.2),
        ("Abertura do malote",       "m_ab_malote", 4.2),
        ("Distribuição dos pacotes", "m_distrib",   4.2),
    ])
    y = row_fields(c, y, [
        ("Fechamento do portão",  "m_fch_portao",   4.2),
        ("Início da aplicação",   "m_inicio",        4.2),
        ("Encerramento",          "m_encerramento",  4.2),
        ("Fechamento do malote",  "m_fch_malote",    4.2),
    ])
    y -= 0.1*cm

    # ── Turno Manhã — Testemunhas ──────────────────────────────────
    y = section_bar(c, y, "3. Turno Manhã — Testemunhas")
    y -= 0.3*cm
    y = witness_table(c, y, "m_port", "Testemunhas — Abertura do Portão")
    y = witness_table(c, y, "m_mal",  "Testemunhas — Abertura do Malote")

    # ── Estatística Manhã (mesma página) ──────────────────────────
    y = section_bar(c, y, "4. Turno Manhã — Estatística por Cargo")
    y -= 0.3*cm
    y = cargo_table(c, y, "m", cargo_data=cargo_data.get("manha") if cargo_data else None)

    page_footer(c, 1)
    c.showPage()

    # ═══════════════════════════════════════════════════════════════
    # PÁGINA 2 — Ocorrências Manhã + início Turno Tarde
    # ═══════════════════════════════════════════════════════════════

    y = draw_main_header(c, subtitle="Turno Manhã — Ocorrências  /  Turno Tarde")

    # ── Ocorrências Manhã ─────────────────────────────────────────
    y = section_bar(c, y, "5. Turno Manhã — Ocorrências")
    y -= 0.4*cm

    y = occurrence_item(c, y, "Ocorrências em sala",           "m_sala",  h_desc=1.5*cm, n_candidates=3)
    y = occurrence_item(c, y, "Prova condicional",             "m_cond",  h_desc=1.5*cm, n_candidates=3)
    y = occurrence_item(c, y, "Toque de celular",              "m_cel",   h_desc=0.8*cm, n_candidates=3)
    y = occurrence_item(c, y, "Declaração de comparecimento",  "m_decl",  h_desc=0.7*cm)

    page_footer(c, 2)
    c.showPage()

    # ═══════════════════════════════════════════════════════════════
    # PÁGINA 3 — Turno Tarde completo + Observações + Assinatura
    # ═══════════════════════════════════════════════════════════════

    y = draw_main_header(c, subtitle="Turno Tarde — Horários, Testemunhas e Estatística")

    # ── Tarde — Horários ──────────────────────────────────────────
    y = section_bar(c, y, "6. Turno Tarde — Horários", color=VERDE)
    y -= 0.45*cm

    y = row_fields(c, y, [
        ("Chegada da equipe",        "t_chegada",   4.2),
        ("Abertura do portão",       "t_ab_portao", 4.2),
        ("Abertura do malote",       "t_ab_malote", 4.2),
        ("Distribuição dos pacotes", "t_distrib",   4.2),
    ])
    y = row_fields(c, y, [
        ("Fechamento do portão",  "t_fch_portao",   4.2),
        ("Início da aplicação",   "t_inicio",        4.2),
        ("Encerramento",          "t_encerramento",  4.2),
        ("Fechamento do malote",  "t_fch_malote",    4.2),
    ])
    y -= 0.1*cm

    # ── Tarde — Testemunhas ───────────────────────────────────────
    y = section_bar(c, y, "7. Turno Tarde — Testemunhas", color=VERDE)
    y -= 0.3*cm
    y = witness_table(c, y, "t_port", "Testemunhas — Abertura do Portão")
    y = witness_table(c, y, "t_mal",  "Testemunhas — Abertura do Malote")

    # ── Tarde — Estatística (mesma página) ────────────────────────
    y = section_bar(c, y, "8. Turno Tarde — Estatística por Cargo", color=VERDE)
    y -= 0.3*cm
    y = cargo_table(c, y, "t", cargo_data=cargo_data.get("tarde") if cargo_data else None)

    page_footer(c, 2)
    c.showPage()

    # ═══════════════════════════════════════════════════════════════
    # PÁGINA 3 — Ocorrências Tarde + Observações + Assinatura
    # ═══════════════════════════════════════════════════════════════

    y = draw_main_header(c, subtitle="Turno Tarde — Ocorrências, Observações e Assinatura")

    # ── Tarde — Ocorrências ───────────────────────────────────────
    y = section_bar(c, y, "9. Turno Tarde — Ocorrências", color=VERDE)
    y -= 0.4*cm

    y = occurrence_item(c, y, "Ocorrências em sala",           "t_sala",  h_desc=1.5*cm, n_candidates=3)
    y = occurrence_item(c, y, "Prova condicional",             "t_cond",  h_desc=1.5*cm, n_candidates=3)
    y = occurrence_item(c, y, "Toque de celular",              "t_cel",   h_desc=0.8*cm, n_candidates=3)
    y = occurrence_item(c, y, "Declaração de comparecimento",  "t_decl",  h_desc=0.7*cm)

    # ── Observações gerais + Assinatura ───────────────────────────
    if y > 3.5*cm:
        y = section_bar(c, y, "10. Observações Gerais")
        y -= 0.35*cm
        obs_h = max(1.5*cm, min(y - 2.8*cm, 2.5*cm))
        field_bg(c, MARGIN, y - obs_h, INNER_W, obs_h)
        txt(c, MARGIN, y - obs_h, INNER_W, obs_h, "observacoes", multiline=True)
        y -= obs_h + 0.5*cm

        c.setStrokeColor(CINZA_LN); c.setLineWidth(0.8)
        c.line(MARGIN, y, MARGIN + 8*cm, y)
        lbl(c, MARGIN, y - 0.35*cm, "Assinatura do Coordenador Local")
        c.line(MARGIN + 10*cm, y, W - MARGIN, y)
        lbl(c, MARGIN + 10*cm, y - 0.35*cm, "Data")

    page_footer(c, 3, total=4)

    # ═══════════════════════════════════════════════════════════════
    # PÁGINA 4 — Registro Fotográfico
    # ═══════════════════════════════════════════════════════════════
    c.showPage()
    y = draw_main_header(c, subtitle="Registro Fotográfico")

    y = section_bar(c, y, "11. Registro Fotográfico")
    y -= 0.4*cm

    lbl(c, MARGIN, y,
        "Insira abaixo as fotos referentes à aplicação. No sistema online, faça upload direto. "
        "Em PDF: abra no Adobe Acrobat → clique na área → insira imagem.",
        col=CINZA, size=7)
    y -= 0.55*cm

    foto_labels = [
        ("foto_sinalizacao",  "Sinalização da Escola",
         "Foto da fachada/identificação do local de prova"),
        ("foto_ab_malote",    "Abertura do Malote",
         "Foto do momento de abertura do malote com testemunhas"),
        ("foto_andamento",    "Andamento da Aplicação",
         "Foto da sala durante a aplicação (candidatos / organização)"),
        ("foto_fch_malote",   "Fechamento do Malote",
         "Foto do malote lacrado com testemunhas ao final"),
    ]

    col_gap  = 0.5*cm
    foto_w   = (INNER_W - col_gap) / 2
    foto_h   = 8.2*cm
    label_h  = 0.9*cm

    positions = [
        (MARGIN,                 y - foto_h - label_h),
        (MARGIN + foto_w + col_gap, y - foto_h - label_h),
        (MARGIN,                 y - (foto_h + label_h) * 2 - 0.55*cm),
        (MARGIN + foto_w + col_gap, y - (foto_h + label_h) * 2 - 0.55*cm),
    ]

    for idx, ((name, title, hint), (px, py)) in enumerate(zip(foto_labels, positions)):
        # Cabeçalho da área
        c.setFillColor(AZUL_MED)
        c.rect(px, py + foto_h, foto_w, label_h, fill=1, stroke=0)
        c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 8)
        c.drawString(px + 0.2*cm, py + foto_h + 0.55*cm, title)
        c.setFont("Helvetica", 6.5)
        c.drawString(px + 0.2*cm, py + foto_h + 0.2*cm, hint)

        # Área da foto — borda tracejada
        c.setStrokeColor(colors.HexColor("#BBBBBB")); c.setLineWidth(0.8)
        c.setDash(5, 4)
        c.rect(px, py, foto_w, foto_h, fill=0, stroke=1)
        c.setDash()

        # Ícone e texto central
        c.setFillColor(colors.HexColor("#DDDDDD"))
        c.rect(px + foto_w/2 - 1.0*cm, py + foto_h/2 + 0.1*cm, 2.0*cm, 1.4*cm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#AAAAAA")); c.setFont("Helvetica", 7)
        c.drawCentredString(px + foto_w/2, py + foto_h/2 - 0.25*cm, "[ inserir foto ]")

        # Campo de legenda abaixo da foto
        lbl(c, px, py - 0.22*cm, "Legenda / descrição:", col=CINZA, size=6.5)
        field_bg(c, px, py - 0.22*cm - FH, foto_w, FH)
        txt(c, px, py - 0.22*cm - FH, foto_w, FH, f"{name}_legenda")

    page_footer(c, 4, total=4)
    c.save()


# ─────────────────────────────────────────────────────────────────
# Batch generation from Excel
# ─────────────────────────────────────────────────────────────────

def sanitize(name):
    return re.sub(r'[\\/*?:"<>|]', '_', str(name)).strip()


# ─────────────────────────────────────────────────────────────────
# Parsing da planilha de alocação
# ─────────────────────────────────────────────────────────────────

def parse_alocacao(xlsx_path):
    """
    Lê a planilha de alocação (aba 'ALOCAÇÃO POR SALA').
    Retorna dict: escola_upper → {
        "manha": [(sala, cargo, previstos), ...],
        "tarde": [(sala, cargo, previstos), ...]
    }
    Mantém a hierarquia Sala → Cargo para exibir na tabela.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'ALOCAÇÃO POR SALA' not in wb.sheetnames:
        return {}
    ws = wb['ALOCAÇÃO POR SALA']

    def turno_key(dt):
        return "manha" if (dt.hour * 60 + dt.minute) < 720 else "tarde"

    result = {}      # escola_upper → {"manha": [], "tarde": []}
    current_school  = None
    current_turno   = "manha"
    current_sala    = None
    SKIP = {"total geral", "total"}

    for row in ws.iter_rows(min_row=2, values_only=True):
        val, count = row[0], row[1]
        if val is None:
            continue

        # Marcador de turno
        if isinstance(val, datetime):
            current_turno  = turno_key(val)
            current_school = None
            current_sala   = None
            continue

        val = str(val).strip()
        if not val:
            continue
        count_int = int(count) if isinstance(count, (int, float)) and count else 0
        val_lower = val.lower()

        if val_lower in SKIP or val_lower.startswith("total"):
            continue

        # Escola
        if count_int > 10 and not val.upper().startswith("SALA") and not val[0].isdigit():
            current_school = val.upper()
            current_sala   = None
            if current_school not in result:
                result[current_school] = {"manha": [], "tarde": []}

        # Sala
        elif current_school and val.upper().startswith("SALA"):
            current_sala = val

        # Cargo
        elif current_school and current_sala and val and val[0].isdigit() and count_int > 0:
            result[current_school][current_turno].append(
                (current_sala, val, count_int)
            )

    return result


def fuzzy_match_escola(escola_name, alocacao_dict):
    """
    Encontra a escola mais similar no dicionário de alocação.
    Retorna (matched_key, score) ou (None, 0) se score < 0.4.
    """
    escola_up = escola_name.upper()
    best_key, best_score = None, 0
    for key in alocacao_dict:
        score = SequenceMatcher(None, escola_up, key).ratio()
        # Bonus se palavras-chave coincidirem
        words_e = set(re.split(r'\W+', escola_up))
        words_k = set(re.split(r'\W+', key))
        overlap = len(words_e & words_k) / max(len(words_e | words_k), 1)
        combined = score * 0.5 + overlap * 0.5
        if combined > best_score:
            best_score = combined
            best_key = key
    return (best_key, best_score) if best_score >= 0.35 else (None, 0)

def batch_from_excel(xlsx_path, escola_filter=None, alocacao_path=None):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, name):
        aliases = {
            "concurso":      ["concurso", "processo", "nome do concurso"],
            "edital":        ["edital", "nº edital", "numero edital"],
            "data_aplicacao":["data", "data aplicação", "data da aplicação", "data_aplicacao"],
            "uf":            ["uf", "estado"],
            "municipio":     ["municipio", "município", "cidade"],
            "local_escola":  ["escola", "local", "local/escola", "local da prova", "nome da escola"],
            "coord_ped":     ["coord. pedagógico", "coordenador pedagógico", "coord_ped", "coord pedagogico"],
            "coord_log":     ["coord. logístico", "coordenador logístico", "coord_log", "coord logistico"],
            "coord_local":   ["coordenação local", "coord. local", "coord_local", "coordenacao local"],
            "turno":         ["turno"],
        }
        for key, possible in aliases.items():
            if name == key:
                for p in possible:
                    if p in headers:
                        idx = headers.index(p)
                        val = row[idx].value
                        return str(val).strip() if val is not None else ""
        return ""

    alocacao_dict = parse_alocacao(alocacao_path) if alocacao_path else {}

    outdir = os.path.dirname(xlsx_path) or "."
    zip_path = os.path.join(outdir, "formularios_prova.zip")

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        count = 0
        for row in ws.iter_rows(min_row=2):
            escola = col(row, "local_escola")
            if not escola:
                continue
            if escola_filter and escola_filter.lower() not in escola.lower():
                continue

            data = {k: col(row, k) for k in
                    ["concurso","edital","data_aplicacao","uf","municipio",
                     "local_escola","coord_ped","coord_log","coord_local","turno"]}

            # Busca cargos por turno na planilha de alocação (se fornecida)
            escola_cargos = None
            if alocacao_dict:
                matched, score = fuzzy_match_escola(escola, alocacao_dict)
                if matched:
                    escola_cargos = alocacao_dict[matched]  # {"manha": [...], "tarde": [...]}
                    nm = len(escola_cargos.get("manha", []))
                    nt = len(escola_cargos.get("tarde", []))
                    print(f"     ↳ {matched} (score={score:.2f}) — Manhã: {nm} cargos, Tarde: {nt} cargos")
                else:
                    print(f"     ↳ Sem correspondência na alocação para '{escola}'")

            fname = sanitize(f"Formulario_{escola}") + ".pdf"
            tmp_path = os.path.join(outdir, fname)
            generate_pdf(tmp_path, data, cargo_data=escola_cargos)
            apply_number_restrictions(tmp_path)
            zf.write(tmp_path, fname)
            os.remove(tmp_path)
            count += 1
            print(f"  ✔  {fname}")

    print(f"\n{count} formulário(s) gerado(s) → {zip_path}")
    return zip_path


# ─────────────────────────────────────────────────────────────────
# Pós-processamento: restrição numérica em campos de hora e CPF
# ─────────────────────────────────────────────────────────────────

def apply_number_restrictions(pdf_path):
    """
    Adiciona validação JavaScript (keystroke) nos campos de horário e CPF,
    aceitando apenas dígitos (e ':' para horários).
    """
    from pypdf import PdfReader, PdfWriter
    from pypdf.generic import (
        DictionaryObject, NameObject, DecodedStreamObject, ArrayObject,
        IndirectObject
    )

    # Campos de horário: aceita dígitos e ":"
    TIME_FIELDS = {
        "m_chegada","m_ab_portao","m_fch_portao","m_distrib",
        "m_inicio","m_encerramento","m_fch_malote","m_ab_malote",
        "t_chegada","t_ab_portao","t_fch_portao","t_distrib",
        "t_inicio","t_encerramento","t_fch_malote","t_ab_malote",
    }
    # Campos de CPF: só dígitos
    CPF_PATTERN = re.compile(r".*_cpf_\d+$")

    def make_js_action(js_code):
        action = DictionaryObject()
        action[NameObject("/S")] = NameObject("/JavaScript")
        stream = DecodedStreamObject()
        stream.set_data(js_code.encode())
        action[NameObject("/JS")] = stream
        return action

    JS_TIME = "event.rc = /^[\\d:]*$/.test(event.change);"
    JS_NUMS = "event.rc = /^\\d*$/.test(event.change);"

    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    writer.clone_reader_document_root(reader)

    modified = 0
    if "/AcroForm" in writer._root_object and "/Fields" in writer._root_object["/AcroForm"]:
        fields = writer._root_object["/AcroForm"]["/Fields"]

        def process_fields(field_list):
            nonlocal modified
            for ref in field_list:
                try:
                    obj = ref.get_object() if hasattr(ref, 'get_object') else ref
                    if not isinstance(obj, DictionaryObject):
                        continue
                    name = obj.get("/T", "")
                    if hasattr(name, 'decode'):
                        name = name.decode('utf-8', errors='ignore')
                    else:
                        name = str(name).strip("()")

                    js = None
                    if name in TIME_FIELDS:
                        js = JS_TIME
                    elif CPF_PATTERN.match(name):
                        js = JS_NUMS

                    if js:
                        aa = DictionaryObject()
                        aa[NameObject("/K")] = make_js_action(js)
                        obj[NameObject("/AA")] = aa
                        modified += 1

                    if "/Kids" in obj:
                        process_fields(obj["/Kids"])
                except Exception:
                    pass

        process_fields(fields)

    writer.write(pdf_path)
    if modified:
        print(f"  → Restrição numérica aplicada em {modified} campo(s)")


# ─────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) == 1:
        out = "/sessions/epic-determined-dijkstra/mnt/outputs/formulario_aplicacao_prova.pdf"
        generate_pdf(out)
        apply_number_restrictions(out)
        print("Template em branco gerado:", out)
    elif len(sys.argv) >= 2:
        xlsx = sys.argv[1]
        escola_filter = sys.argv[2] if len(sys.argv) >= 3 else None
        batch_from_excel(xlsx, escola_filter)
