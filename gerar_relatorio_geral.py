"""
Gerador do Relatório Geral de Aplicação de Prova — IBGP
Consolida os dados de todas as escolas em PDF + Word.

Uso:
  python gerar_relatorio_geral.py                        → gera com dados de exemplo
  python gerar_relatorio_geral.py planilha_ibgp.xlsx     → lê dados reais da planilha
  python gerar_relatorio_geral.py planilha_ibgp.xlsx saida/  → define pasta de saída
"""

import os, sys, random
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

W, H = A4

# ── Paleta IBGP ───────────────────────────────────────────────────
VERMELHO   = colors.HexColor("#CF3432")
CINZA_ESC  = colors.HexColor("#393939")
CINZA_MED  = colors.HexColor("#4A4A4A")
CINZA_CL   = colors.HexColor("#F4F4F4")
CINZA_LN   = colors.HexColor("#CCCCCC")
BRANCO     = colors.white
CINZA_TXT  = colors.HexColor("#555555")

MARGIN     = 1.5*cm
INNER_W    = W - 2 * MARGIN

# Logo: procura na pasta do script, depois em logo/, depois em uploads/
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
for _lp in [
    os.path.join(_SCRIPT_DIR, "IBGP.png"),
    os.path.join(_SCRIPT_DIR, "logo", "IBGP.png"),
    "/sessions/epic-determined-dijkstra/mnt/uploads/IBGP.png",
]:
    if os.path.exists(_lp):
        LOGO_PATH = _lp
        break
else:
    LOGO_PATH = os.path.join(_SCRIPT_DIR, "IBGP.png")  # fallback


# ─────────────────────────────────────────────────────────────────
# DADOS SIMULADOS (modelo)
# ─────────────────────────────────────────────────────────────────

CONCURSO_INFO = {
    "titulo":            "Concurso Público do Município de Alto Rio Doce / MG",
    "edital":            "Edital nº 01/2026",
    "data":              "26 de julho de 2026",
    "turno":             "Domingo Manhã",
    "coord_ped":         "Tiago Theisen",
    "coord_log":         "Marcela Avelar",
    "gerado_em":         date.today().strftime("%d/%m/%Y"),
    "num_locais_extenso":"quatro",
    # Horários padronizados da aplicação
    "hr_reuniao_ini":    "05h50",
    "hr_reuniao_fim":    "06h45",
    "hr_ab_portao":      "07h00",
    "hr_distrib":        "07h30",
    "hr_fch_portao":     "08h00",
    "hr_inicio_prova":   "08h30",
}

# Dados consolidados por escola — manhã e tarde
# Tupla: (sala, cargo, inscritos, presentes, ausentes, inclusao)
ESCOLAS = [
    {
        "nome":          "Secretaria Municipal de Educação - ARD",
        "coord_local":   "Natália Cristina Matias",
        "turno":         "AMBOS",
        "salas":         5,
        "andamento":     "Tranquilo, sem intercorrências relevantes",
        "toque_celular": "Não consta",
        "ocorrencias":   "Nenhuma ocorrência registrada.",
        "manha": [
            ("Sala 01", "01 Agente de Combate a Endemias",      12, 10, 2, 0),
            ("Sala 01", "01 Agente de Controle de Vetores",      8,  7, 1, 0),
            ("Sala 02", "01 Assistente Administrativo",         15, 13, 2, 1),
            ("Sala 02", "01 Assistente Social",                  6,  5, 1, 0),
            ("Sala 03", "01 Auxiliar de Serviços Gerais",       18, 15, 3, 0),
        ],
        "tarde": [
            ("Sala 04", "01 Enfermeiro",                        10,  9, 1, 0),
            ("Sala 04", "01 Farmacêutico",                       5,  4, 1, 0),
            ("Sala 05", "01 Fisioterapeuta",                     7,  7, 0, 0),
            ("Sala 05", "01 Médico Clínico Geral",               4,  3, 1, 0),
        ],
    },
    {
        "nome":          "Escola Municipal Aristides da Mota Marinho - ARD",
        "coord_local":   "Jordan Guilherme",
        "turno":         "MANHÃ",
        "salas":         3,
        "andamento":     "Tranquilo, sem intercorrências relevantes",
        "toque_celular": "Sim",
        "ocorrencias":   "Toque de celular em sala — Sala 01, candidato orientado e advertido.",
        "manha": [
            ("Sala 01", "01 Contador",                           9,  8, 1, 0),
            ("Sala 01", "01 Educador Físico",                    6,  6, 0, 0),
            ("Sala 02", "01 Engenheiro Civil",                   5,  4, 1, 0),
            ("Sala 02", "01 Fonoaudiólogo",                      4,  4, 0, 0),
            ("Sala 03", "01 Médico Pediatra",                    3,  3, 0, 0),
        ],
        "tarde": [],
    },
    {
        "nome":          "Escola Municipal Raul Soares - ARD",
        "coord_local":   "Kênia Soiane",
        "turno":         "AMBOS",
        "salas":         4,
        "andamento":     "Tranquilo, com registro de prova condicional",
        "toque_celular": "Não consta",
        "ocorrencias":   "Candidato com prova condicional — Sala 02 (manhã). Documentação encaminhada ao malote.",
        "manha": [
            ("Sala 01", "01 Nutricionista",                      7,  6, 1, 0),
            ("Sala 01", "01 Odontólogo",                         8,  7, 1, 0),
            ("Sala 02", "01 Operador de Máquinas",              14, 12, 2, 1),
            ("Sala 02", "01 Pedreiro",                          16, 14, 2, 0),
        ],
        "tarde": [
            ("Sala 03", "01 Psicólogo",                          6,  5, 1, 0),
            ("Sala 03", "01 Técnico em Enfermagem",             12, 11, 1, 0),
            ("Sala 04", "01 Técnico em Informática",             8,  7, 1, 1),
        ],
    },
    {
        "nome":          "UNIPAC — Campus Bonifácio de Andrada - Barbacena",
        "coord_local":   "Laryssa Fugêncio",
        "turno":         "AMBOS",
        "salas":         5,
        "andamento":     "Tranquilo, com ocorrência em sala registrada no turno tarde",
        "toque_celular": "Sim",
        "ocorrencias":   "Ocorrência em sala — Sala 03 (tarde): candidato retirado por utilização de dispositivo eletrônico.",
        "manha": [
            ("Sala 01", "01 Auxiliar Administrativo",           18, 16, 2, 0),
            ("Sala 01", "01 Auxiliar de Biblioteca",            10,  9, 1, 0),
            ("Sala 02", "01 Auxiliar de Farmácia",              12, 10, 2, 1),
            ("Sala 02", "01 Cozinheiro",                        15, 13, 2, 0),
            ("Sala 03", "01 Gari",                              20, 17, 3, 0),
            ("Sala 03", "01 Mecânico",                           8,  7, 1, 0),
        ],
        "tarde": [
            ("Sala 04", "01 Motorista",                         11, 10, 1, 0),
            ("Sala 04", "01 Operador de Escavadeira",            6,  5, 1, 0),
            ("Sala 05", "01 Recepcionista",                     13, 12, 1, 0),
            ("Sala 05", "01 Vigilante",                         16, 14, 2, 1),
        ],
    },
]


# ─────────────────────────────────────────────────────────────────
# HELPERS PDF
# ─────────────────────────────────────────────────────────────────

def draw_header(c, subtitle=None):
    hdr_h = 2.8*cm
    c.setFillColor(VERMELHO)
    c.rect(0, H - 0.28*cm, W, 0.28*cm, fill=1, stroke=0)
    try:
        logo = ImageReader(LOGO_PATH)
        logo_w, logo_h = 5.2*cm, 5.2*cm * (711/2200)
        logo_y = H - hdr_h + (hdr_h - 0.28*cm - logo_h) / 2
        c.drawImage(logo, MARGIN, logo_y, width=logo_w, height=logo_h, mask='auto')
    except Exception:
        c.setFillColor(CINZA_ESC); c.setFont("Helvetica-Bold", 14)
        c.drawString(MARGIN, H - 1.6*cm, "IBGP")
    div_x = MARGIN + 5.2*cm + 0.6*cm
    c.setStrokeColor(VERMELHO); c.setLineWidth(1.2)
    c.line(div_x, H - 0.45*cm, div_x, H - hdr_h + 0.3*cm)
    txt_x = div_x + 0.45*cm
    c.setFillColor(CINZA_ESC); c.setFont("Helvetica-Bold", 12)
    c.drawString(txt_x, H - 1.2*cm, "RELATÓRIO GERAL DE APLICAÇÃO DE PROVA")
    if subtitle:
        c.setFillColor(VERMELHO); c.setFont("Helvetica-Bold", 8)
        c.drawString(txt_x, H - 1.7*cm, subtitle)
    c.setFillColor(CINZA_TXT); c.setFont("Helvetica", 7.5)
    c.drawString(txt_x, H - 2.1*cm, "Documento consolidado — uso interno IBGP")
    c.setStrokeColor(VERMELHO); c.setLineWidth(1.5)
    c.line(0, H - hdr_h, W, H - hdr_h)
    return H - hdr_h - 0.3*cm


def draw_footer(c, page_num, total):
    c.setFillColor(VERMELHO)
    c.rect(0, 0.6*cm, W, 0.12*cm, fill=1, stroke=0)
    c.setFillColor(CINZA_ESC)
    c.rect(0, 0, W, 0.6*cm, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica", 6.5)
    c.drawCentredString(W/2, 0.2*cm,
        "Relatório Geral de Aplicação de Prova  •  IBGP — Instituto Brasileiro de Gestão e Pesquisa  •  Uso interno")
    c.drawRightString(W - MARGIN, 0.2*cm, f"Pág. {page_num}/{total}")


def section_bar(c, y, title, color=CINZA_ESC, h=0.58*cm):
    c.setFillColor(color)
    c.rect(MARGIN, y - h, INNER_W, h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 8.5)
    c.drawString(MARGIN + 0.3*cm, y - h + 0.17*cm, title.upper())
    return y - h


def lbl(c, x, y, text, bold=False, size=7.5, col=CINZA_TXT):
    c.setFillColor(col)
    c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
    c.drawString(x, y, text)


def draw_table_row(c, y, cells, col_xs, col_ws, row_h, bg, text_size=7.5, bold_cols=None):
    bold_cols = bold_cols or []
    for i, (text, cx, cw) in enumerate(zip(cells, col_xs, col_ws)):
        c.setFillColor(bg)
        c.setStrokeColor(CINZA_LN); c.setLineWidth(0.3)
        c.rect(cx, y - row_h, cw, row_h, fill=1, stroke=0)
        c.setFillColor(CINZA_ESC if i in bold_cols else CINZA_TXT)
        c.setFont("Helvetica-Bold" if i in bold_cols else "Helvetica", text_size)
        # clip
        max_chars = int(cw / (0.135*cm)) + 1
        display = str(text)[:max_chars]
        c.drawString(cx + 0.12*cm, y - row_h + (row_h - text_size/72*cm)/2 + 0.05*cm, display)
    # linha inferior
    c.setStrokeColor(CINZA_LN); c.setLineWidth(0.25)
    c.line(col_xs[0], y - row_h, col_xs[-1] + col_ws[-1], y - row_h)


def draw_table_header(c, y, headers, col_xs, col_ws, row_h=0.38*cm):
    c.setFillColor(CINZA_MED)
    c.rect(col_xs[0], y - row_h, sum(col_ws), row_h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7)
    for text, cx, cw in zip(headers, col_xs, col_ws):
        c.drawString(cx + 0.12*cm, y - row_h + 0.1*cm, text)
    return y - row_h


def table_border(c, y_top, y_bot, col_xs, col_ws):
    total_w = col_xs[-1] + col_ws[-1] - col_xs[0]
    c.setStrokeColor(CINZA_LN); c.setLineWidth(0.5)
    c.rect(col_xs[0], y_bot, total_w, y_top - y_bot, fill=0, stroke=1)
    # separadores verticais
    for i in range(1, len(col_xs)):
        c.line(col_xs[i], y_top, col_xs[i], y_bot)


# ─────────────────────────────────────────────────────────────────
# GERADOR PDF
# ─────────────────────────────────────────────────────────────────

def gerar_pdf(output_path, info, escolas):
    c = canvas.Canvas(output_path, pagesize=A4)
    c.setTitle("Relatório Geral de Aplicação de Prova")

    TOTAL_PAGES = 3  # estimado; atualiza se necessário

    # ══════════════════════════════════════════════
    # PÁGINA 1 — Capa + Resumo Identificação
    # ══════════════════════════════════════════════
    y = draw_header(c)

    # Bloco de identificação do concurso
    y = section_bar(c, y, "Identificação do Concurso")
    y -= 0.5*cm

    id_items = [
        ("Concurso:",             info["titulo"]),
        ("Edital:",               info["edital"]),
        ("Data da Aplicação:",    info["data"]),
        ("Coord. Pedagógico:",    info["coord_ped"]),
        ("Coord. Logístico:",     info["coord_log"]),
        ("Total de Locais:",      str(len(escolas))),
        ("Relatório gerado em:",  info["gerado_em"]),
    ]
    for lbl_text, val_text in id_items:
        lbl(c, MARGIN, y, lbl_text, bold=True, size=8, col=CINZA_MED)
        lbl(c, MARGIN + 5.5*cm, y, val_text, size=8)
        y -= 0.5*cm
    y -= 0.1*cm

    # ── Resumo consolidado ────────────────────────
    y = section_bar(c, y, "1. Resumo Consolidado — Totais Gerais")
    y -= 0.45*cm

    # Calcular totais gerais
    tot_prev = tot_pres = tot_aus = tot_inc = 0
    for e in escolas:
        for turno in ("manha", "tarde"):
            for row in e.get(turno, []):
                _, _, prev, pres, aus, inc = row
                tot_prev += prev; tot_pres += pres
                tot_aus  += aus;  tot_inc  += inc

    # Cards de totais
    card_labels = ["PREVISTOS", "PRESENTES", "AUSENTES", "INCLUSÃO"]
    card_values = [tot_prev, tot_pres, tot_aus, tot_inc]
    card_colors = [CINZA_ESC, colors.HexColor("#2E7D32"),
                   VERMELHO,  colors.HexColor("#0D47A1")]
    card_w = (INNER_W - 0.9*cm) / 4
    card_h = 2.0*cm
    cx_start = MARGIN

    for i, (lbl_t, val, col) in enumerate(zip(card_labels, card_values, card_colors)):
        cx = cx_start + i * (card_w + 0.3*cm)
        # Card fundo
        c.setFillColor(col)
        c.setStrokeColor(BRANCO); c.setLineWidth(0)
        c.rect(cx, y - card_h, card_w, card_h, fill=1, stroke=0)
        # Valor grande
        c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 22)
        c.drawCentredString(cx + card_w/2, y - card_h + 0.7*cm, str(val))
        # Label
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(cx + card_w/2, y - card_h + 0.22*cm, lbl_t)

    y -= card_h + 0.5*cm

    # Percentual de presença
    pct = round(tot_pres / tot_prev * 100, 1) if tot_prev else 0
    lbl(c, MARGIN, y, f"Índice de presença geral: {pct}%  |  "
        f"Ausência: {round(100-pct,1)}%", bold=True, size=9, col=CINZA_ESC)
    y -= 0.6*cm

    # ── Tabela por escola ──────────────────────────
    y = section_bar(c, y, "2. Resultado por Escola")
    y -= 0.35*cm

    col_xs = [MARGIN, MARGIN+7.5*cm, MARGIN+9.3*cm, MARGIN+11.1*cm, MARGIN+12.9*cm, MARGIN+14.7*cm]
    col_ws = [7.3*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.6*cm, INNER_W-(col_xs[-1]-MARGIN)]
    headers_esc = ["Escola / Local", "Turno", "Previstos", "Presentes", "Ausentes", "Inclusão"]

    y_top_esc = y
    y = draw_table_header(c, y, headers_esc, col_xs, col_ws)
    row_h = 0.45*cm

    for i, e in enumerate(escolas):
        bg = CINZA_CL if i % 2 == 0 else colors.HexColor("#EBEBEB")
        e_prev = sum(r[2] for t in ("manha","tarde") for r in e.get(t,[]))
        e_pres = sum(r[3] for t in ("manha","tarde") for r in e.get(t,[]))
        e_aus  = sum(r[4] for t in ("manha","tarde") for r in e.get(t,[]))
        e_inc  = sum(r[5] for t in ("manha","tarde") for r in e.get(t,[]))
        draw_table_row(c, y, [e["nome"], e["turno"], e_prev, e_pres, e_aus, e_inc],
                       col_xs, col_ws, row_h, bg, text_size=7)
        y -= row_h

    # Linha de totais
    c.setFillColor(CINZA_ESC)
    c.rect(col_xs[0], y - row_h, sum(col_ws), row_h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7.5)
    c.drawString(col_xs[0]+0.12*cm, y-row_h+(row_h-7.5/72*cm)/2+0.05*cm, "TOTAL GERAL")
    for val, cx, cw in zip([tot_prev,tot_pres,tot_aus,tot_inc], col_xs[2:], col_ws[2:]):
        c.drawString(cx+0.12*cm, y-row_h+(row_h-7.5/72*cm)/2+0.05*cm, str(val))
    y -= row_h

    table_border(c, y_top_esc, y, col_xs, col_ws)
    y -= 0.4*cm

    # ── Ocorrências ────────────────────────────────
    y = section_bar(c, y, "3. Registro de Ocorrências por Escola")
    y -= 0.35*cm

    for e in escolas:
        lbl(c, MARGIN, y, e["nome"], bold=True, size=7.5, col=CINZA_ESC)
        y -= 0.35*cm
        lbl(c, MARGIN + 0.3*cm, y, e.get("ocorrencias","—"), size=7.5)
        y -= 0.45*cm

    draw_footer(c, 1, TOTAL_PAGES)
    c.showPage()

    # ══════════════════════════════════════════════
    # PÁGINA 2 — Estatística Detalhada por Cargo
    # ══════════════════════════════════════════════
    y = draw_header(c, subtitle="Estatística Detalhada por Cargo")

    y = section_bar(c, y, "4. Estatística por Cargo — Todos os Locais")
    y -= 0.35*cm

    # Agregar por cargo (nome do cargo sem prefixo numérico)
    from collections import defaultdict
    cargo_totals = defaultdict(lambda: [0,0,0,0])  # [prev, pres, aus, inc]
    for e in escolas:
        for turno in ("manha","tarde"):
            for row in e.get(turno, []):
                _, cargo, prev, pres, aus, inc = row
                cargo_totals[cargo][0] += prev
                cargo_totals[cargo][1] += pres
                cargo_totals[cargo][2] += aus
                cargo_totals[cargo][3] += inc

    col_xs2 = [MARGIN, MARGIN+8.5*cm, MARGIN+10.5*cm, MARGIN+12.5*cm, MARGIN+14.5*cm]
    col_ws2 = [8.3*cm, 1.8*cm, 1.8*cm, 1.8*cm, INNER_W-(col_xs2[-1]-MARGIN)]
    headers2 = ["Cargo / Função", "Previstos", "Presentes", "Ausentes", "Inclusão"]

    y_top2 = y
    y = draw_table_header(c, y, headers2, col_xs2, col_ws2)
    row_h = 0.42*cm

    g_prev = g_pres = g_aus = g_inc = 0
    for i, (cargo, totais) in enumerate(sorted(cargo_totals.items())):
        bg = CINZA_CL if i % 2 == 0 else colors.HexColor("#EBEBEB")
        prev, pres, aus, inc = totais
        g_prev+=prev; g_pres+=pres; g_aus+=aus; g_inc+=inc
        draw_table_row(c, y, [cargo, prev, pres, aus, inc],
                       col_xs2, col_ws2, row_h, bg, text_size=7)
        y -= row_h
        if y < 2.5*cm:
            draw_footer(c, 2, TOTAL_PAGES)
            c.showPage()
            y = draw_header(c, subtitle="Estatística por Cargo (continuação)")
            y_top2 = y
            y = draw_table_header(c, y, headers2, col_xs2, col_ws2)

    # Totais
    c.setFillColor(CINZA_ESC)
    c.rect(col_xs2[0], y-row_h, sum(col_ws2), row_h, fill=1, stroke=0)
    c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7.5)
    c.drawString(col_xs2[0]+0.12*cm, y-row_h+(row_h-7.5/72*cm)/2+0.05*cm, "TOTAL GERAL")
    for val, cx in zip([g_prev, g_pres, g_aus, g_inc], col_xs2[1:]):
        c.drawString(cx+0.12*cm, y-row_h+(row_h-7.5/72*cm)/2+0.05*cm, str(val))
    y -= row_h
    table_border(c, y_top2, y, col_xs2, col_ws2)

    draw_footer(c, 2, TOTAL_PAGES)
    c.showPage()

    # ══════════════════════════════════════════════
    # PÁGINA 3 — Detalhamento por Local (Manhã + Tarde)
    # ══════════════════════════════════════════════
    y = draw_header(c, subtitle="Detalhamento por Local de Aplicação")

    col_xs3 = [MARGIN, MARGIN+2.5*cm, MARGIN+8.5*cm, MARGIN+10.4*cm, MARGIN+12.3*cm, MARGIN+14.2*cm]
    col_ws3 = [2.3*cm, 5.8*cm, 1.7*cm, 1.7*cm, 1.7*cm, INNER_W-(col_xs3[-1]-MARGIN)]
    headers3 = ["Sala", "Cargo / Função", "Previstos", "Presentes", "Ausentes", "Inclusão"]
    row_h = 0.40*cm

    page_n = 3
    for e in escolas:
        for turno_key, turno_label, turno_color in [
            ("manha", "TURNO MANHÃ", CINZA_MED),
            ("tarde", "TURNO TARDE", VERMELHO),
        ]:
            rows = e.get(turno_key, [])
            if not rows:
                continue

            needed_h = 0.58*cm + 0.38*cm + len(rows)*row_h + row_h + 0.3*cm
            if y - needed_h < 2.5*cm:
                draw_footer(c, page_n, TOTAL_PAGES)
                c.showPage(); page_n += 1
                y = draw_header(c, subtitle="Detalhamento por Local (continuação)")

            # Título escola + turno
            c.setFillColor(turno_color)
            c.rect(MARGIN, y-0.58*cm, INNER_W, 0.58*cm, fill=1, stroke=0)
            c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 8)
            c.drawString(MARGIN+0.3*cm, y-0.42*cm,
                         f"{e['nome']}  —  {turno_label}  |  Coord. Local: {e['coord_local']}")
            y -= 0.58*cm

            y_top3 = y
            y = draw_table_header(c, y, headers3, col_xs3, col_ws3)

            t_prev = t_pres = t_aus = t_inc = 0
            prev_sala = None
            for i, (sala, cargo, prev, pres, aus, inc) in enumerate(rows):
                bg = CINZA_CL if i % 2 == 0 else colors.HexColor("#EBEBEB")
                sala_display = sala if sala != prev_sala else ""
                draw_table_row(c, y, [sala_display, cargo, prev, pres, aus, inc],
                               col_xs3, col_ws3, row_h, bg, text_size=7)
                t_prev+=prev; t_pres+=pres; t_aus+=aus; t_inc+=inc
                prev_sala = sala
                y -= row_h

            # Subtotal turno
            c.setFillColor(colors.HexColor("#555555"))
            c.rect(col_xs3[0], y-row_h, sum(col_ws3), row_h, fill=1, stroke=0)
            c.setFillColor(BRANCO); c.setFont("Helvetica-Bold", 7)
            c.drawString(col_xs3[0]+0.12*cm, y-row_h+(row_h-7/72*cm)/2+0.05*cm,
                         f"Subtotal {turno_label}")
            for val, cx in zip([t_prev, t_pres, t_aus, t_inc], col_xs3[2:]):
                c.drawString(cx+0.12*cm, y-row_h+(row_h-7/72*cm)/2+0.05*cm, str(val))
            y -= row_h

            table_border(c, y_top3, y, col_xs3, col_ws3)
            y -= 0.5*cm

    draw_footer(c, page_n, TOTAL_PAGES)
    c.save()
    print(f"PDF gerado: {output_path}")


# ─────────────────────────────────────────────────────────────────
# GERADOR WORD (docx via Node.js)
# _DOCX_SCRIPT é reescrito para relatorio_geral_gen.js a cada execução.
# ─────────────────────────────────────────────────────────────────

_DOCX_SCRIPT = r"""











'use strict';
const { Document, Packer, Paragraph, Table, TableRow, TableCell,
        TextRun, ImageRun, AlignmentType, WidthType, ShadingType,
        BorderStyle, LevelFormat, PageBreak } = require('docx');
const fs = require('fs');

// ── Paleta ────────────────────────────────────────────────────────
var VRM   = "CF3432";   // vermelho IBGP
var ESC   = "393939";   // cinza escuro
var MED   = "4A4A4A";
var CL    = "F7F7F7";
var BR    = "FFFFFF";
var BOR   = "CCCCCC";
var LRED  = "FFF0F0";

// Cores indicadores — paleta variada
var IBLUE  = "1F3A6E";   // azul escuro
var IGRN   = "1E6B3B";   // verde
var IRED   = "A52D1E";   // vermelho escuro
var ISLAT  = "2E4B7A";   // azul slate
var ITEAL  = "176B6B";   // teal
var INDIGO = "3B2F72";   // índigo
var DKRED  = "7A1A1A";   // vermelho muito escuro

// ── Env ───────────────────────────────────────────────────────────
var INFO     = JSON.parse(process.env.RELATORIO_INFO);
var DADOS    = JSON.parse(process.env.RELATORIO_DADOS);
var LOGOPATH = process.env.LOGO_PATH || "";

// ── Cálculos ──────────────────────────────────────────────────────
var totInscr=0, totPres=0, totAus=0, totInc=0;
DADOS.forEach(function(e) {
  ['manha','tarde'].forEach(function(t) {
    (e[t]||[]).forEach(function(r){ totInscr+=r[2]; totPres+=r[3]; totAus+=r[4]; totInc+=r[5]; });
  });
});
var pctAus  = totInscr ? (totAus /totInscr*100).toFixed(2) : "0.00";
var pctPres = totInscr ? (totPres/totInscr*100).toFixed(2) : "0.00";
var totalSalas = DADOS.reduce(function(s,e){ return s+e.salas; }, 0);

var cargoMap = {};
DADOS.forEach(function(e) {
  ['manha','tarde'].forEach(function(t) {
    (e[t]||[]).forEach(function(r) {
      if (!cargoMap[r[1]]) cargoMap[r[1]] = {inscr:0,pres:0,aus:0,inc:0};
      cargoMap[r[1]].inscr+=r[2]; cargoMap[r[1]].pres+=r[3];
      cargoMap[r[1]].aus+=r[4];  cargoMap[r[1]].inc+=r[5];
    });
  });
});
var totalCargos = Object.keys(cargoMap).length;

// Turnos efetivamente realizados
var hasManha = DADOS.some(function(e){ return (e.manha||[]).length > 0; });
var haTarde  = DADOS.some(function(e){ return (e.tarde||[]).length > 0; });
var turnoStr = (hasManha && haTarde) ? 'Manhã e Tarde' : hasManha ? 'Manhã' : 'Tarde';

function escolaTot(e) {
  var inscr=0,pres=0,aus=0,inc=0;
  ['manha','tarde'].forEach(function(t) {
    (e[t]||[]).forEach(function(r){ inscr+=r[2]; pres+=r[3]; aus+=r[4]; inc+=r[5]; });
  });
  return {inscr:inscr,pres:pres,aus:aus,inc:inc};
}

function topCargoAus(e) {
  var m = {}; var best=null, bpct=-1;
  ['manha','tarde'].forEach(function(t) {
    (e[t]||[]).forEach(function(r) {
      if (!m[r[1]]) m[r[1]]={inscr:0,aus:0};
      m[r[1]].inscr+=r[2]; m[r[1]].aus+=r[4];
    });
  });
  Object.entries(m).forEach(function(kv) {
    var p = kv[1].inscr ? kv[1].aus/kv[1].inscr*100 : 0;
    if (p>bpct) { bpct=p; best=kv[0]; }
  });
  return best ? best.replace(/^01\s+/,'') + ' (' + bpct.toFixed(0) + '%)' : '—';
}

var escolasT = DADOS.map(function(e){ return Object.assign({},e,escolaTot(e)); });
var top5inscr = escolasT.slice().sort(function(a,b){ return b.inscr-a.inscr; }).slice(0,5);
var top5cargos = Object.entries(cargoMap)
  .map(function(kv){ var d=kv[1];
    return {cargo:kv[0].replace(/^01\s+/,''), inscr:d.inscr, pres:d.pres, aus:d.aus,
            pct: d.inscr?d.aus/d.inscr*100:0}; })
  .sort(function(a,b){ return b.pct-a.pct; }).slice(0,5);

// ── Helpers ───────────────────────────────────────────────────────
var BSINGLE = { style: BorderStyle.SINGLE, size: 4, color: BOR };
var BNONE   = { style: BorderStyle.NONE };

function borders(opts) {
  opts = opts || {};
  var b = opts.none ? BNONE : BSINGLE;
  return { top:b, bottom:b, left:b, right:b };
}

function mkCell(text, opts) {
  opts = opts || {};
  var align = opts.center ? AlignmentType.CENTER : (opts.right ? AlignmentType.RIGHT : AlignmentType.LEFT);
  return new TableCell({
    children: [new Paragraph({
      children: [new TextRun({
        text: String(text==null?'':text),
        bold: opts.bold||false,
        size: opts.sz||24,
        color: opts.col||ESC,
      })],
      alignment: align,
      spacing: { before: 40, after: 40 },
    })],
    shading: { type: ShadingType.CLEAR, fill: opts.bg||BR },
    borders: opts.noBorder ? {top:BNONE,bottom:BNONE,left:BNONE,right:BNONE} : borders(),
    width: { size: opts.w||1000, type: WidthType.DXA },
    margins: { top:60, bottom:60, left:80, right:80 },
    verticalAlign: 'center',
  });
}

function makeTable(headers, rows, colW, opts) {
  opts = opts||{};
  var total = colW.reduce(function(a,b){return a+b;},0);
  var hRow = new TableRow({
    children: headers.map(function(h,i){
      return mkCell(h, {bold:true, col:BR, bg:ESC, w:colW[i], center:i>0&&i!==1, sz:24});
    }),
    tableHeader: true,
  });
  var dRows = rows.map(function(row,ri){
    var isLast = opts.totalRow && ri===rows.length-1;
    var hi = opts.highCol!=null ? parseFloat(row[opts.highCol]) : NaN;
    var isHigh = !isNaN(hi) && hi>=50;
    var bg = isLast ? "E8E8E8" : isHigh ? LRED : (ri%2===0?CL:BR);
    return new TableRow({
      children: row.map(function(v,ci){
        return mkCell(v, {bold:isLast, bg:bg, w:colW[ci], center:ci>0&&ci!==1, sz:24});
      }),
    });
  });
  return new Table({
    rows: [hRow].concat(dRows),
    width: { size:total, type:WidthType.DXA },
    columnWidths: colW,
  });
}

function kvTable(rows, opts) {
  opts = opts||{};
  var w1 = opts.w1||2800, w2 = opts.w2||6400;
  return new Table({
    rows: rows.map(function(r){
      return new TableRow({ children: [
        mkCell(r[0], {bold:true, bg:"F0F0F0", w:w1, sz:24}),
        mkCell(r[1], {w:w2, sz:24}),
      ]});
    }),
    width: { size:w1+w2, type:WidthType.DXA },
    columnWidths: [w1, w2],
  });
}

function space(n) { return new Paragraph({ spacing:{ after: n||100 } }); }
function pgBreak() { return new Paragraph({ children:[new PageBreak()] }); }

function secHead(text, num) {
  return new Paragraph({
    children: [new TextRun({
      text: (num?num+'. ':'')+text,
      bold: true, size: 26, color: ESC,
    })],
    border: { left: { style: BorderStyle.THICK, size: 24, color: VRM } },
    indent: { left: 240 },
    spacing: { before: 320, after: 160 },
  });
}

function subHead(text) {
  return new Paragraph({
    children: [new TextRun({ text:text, bold:true, size:24, color:MED })],
    alignment: AlignmentType.JUSTIFIED,
    spacing: { before:200, after:100 },
  });
}

function para(text, opts) {
  opts=opts||{};
  return new Paragraph({
    children: [new TextRun({ text:text, size:24, color:"444444", italics:opts.italic||false })],
    alignment: opts.center ? AlignmentType.CENTER : AlignmentType.JUSTIFIED,
    spacing: { after: opts.after!=null?opts.after:80 },
  });
}

function bullet(text) {
  return new Paragraph({
    children: [new TextRun({ text:text, size:24, color:ESC})],
    numbering: { reference:"blist", level:0 },
    spacing: { after:50 },
  });
}

// Horários
var HR = {
  reuniao:    (INFO.hr_reuniao_ini||'05h50') + ' às ' + (INFO.hr_reuniao_fim||'06h45'),
  ab_portao:  INFO.hr_ab_portao   || '07h00',
  distrib:    INFO.hr_distrib     || '07h30',
  fch_portao: INFO.hr_fch_portao  || '08h00',
  inicio:     INFO.hr_inicio_prova|| '08h30',
};

// ── Logo ──────────────────────────────────────────────────────────
var logoImg = null;
if (LOGOPATH && fs.existsSync(LOGOPATH)) {
  try { logoImg = fs.readFileSync(LOGOPATH); } catch(e){}
}

// ── CAPA ──────────────────────────────────────────────────────────
var children = [];

// Logo centralizada
children.push(new Paragraph({
  children: logoImg
    ? [new ImageRun({data:logoImg, transformation:{width:240,height:73}, type:"png"})]
    : [new TextRun({text:"IBGP", bold:true, size:64, color:ESC})],
  alignment: AlignmentType.CENTER,
  spacing: {before:800, after:160},
}));

children.push(new Paragraph({
  children:[new TextRun({text:"Instituto Brasileiro de Gestão e Pesquisa", size:22, color:MED})],
  alignment:AlignmentType.CENTER,
  spacing:{after:600},
}));

children.push(space(1000));

// Título principal
children.push(new Paragraph({
  children:[new TextRun({text:"RELATÓRIO TÉCNICO CONSOLIDADO", bold:true, size:52, color:ESC})],
  alignment:AlignmentType.CENTER,
  spacing:{before:0,after:600},
}));

// Nome do concurso
children.push(new Paragraph({
  children:[new TextRun({text:INFO.titulo.toUpperCase(), bold:true, size:34, color:ESC})],
  alignment:AlignmentType.CENTER,
  spacing:{before:0,after:200},
}));

// Edital
children.push(new Paragraph({
  children:[new TextRun({text:INFO.edital.toUpperCase(), bold:true, size:34, color:ESC})],
  alignment:AlignmentType.CENTER,
  spacing:{before:0,after:600},
}));

// Tipo de prova
children.push(new Paragraph({
  children:[new TextRun({text:"PROVA OBJETIVA", bold:true, size:32, color:MED})],
  alignment:AlignmentType.CENTER,
  spacing:{before:0,after:1400},
}));

// Informações — parágrafos simples, à esquerda
children.push(new Paragraph({
  children:[
    new TextRun({text:"Data: ", size:24, color:MED}),
    new TextRun({text:INFO.data, bold:true, size:24, color:ESC}),
  ],
  spacing:{after:180},
}));

children.push(new Paragraph({
  children:[
    new TextRun({text:"Turno: ", size:24, color:MED}),
    new TextRun({text:turnoStr, bold:true, size:24, color:ESC}),
  ],
  spacing:{after:180},
}));

children.push(new Paragraph({
  children:[
    new TextRun({text:"Total de locais: ", size:24, color:MED}),
    new TextRun({text:String(DADOS.length), bold:true, size:24, color:ESC}),
  ],
  spacing:{after:600},
}));

children.push(new Paragraph({
  children:[
    new TextRun({text:"Coordenador Pedagógico: ", size:24, color:MED}),
    new TextRun({text:INFO.coord_ped, bold:true, size:24, color:ESC}),
  ],
  alignment:AlignmentType.RIGHT,
  spacing:{after:180},
}));

children.push(new Paragraph({
  children:[
    new TextRun({text:"Coordenador Logístico: ", size:24, color:MED}),
    new TextRun({text:INFO.coord_log, bold:true, size:24, color:ESC}),
  ],
  alignment:AlignmentType.RIGHT,
  spacing:{after:180},
}));

children.push(new Paragraph({
  children:[
    new TextRun({text:"Gerado em: ", size:24, color:MED}),
    new TextRun({text:INFO.gerado_em, bold:true, size:24, color:ESC}),
  ],
  alignment:AlignmentType.RIGHT,
  spacing:{after:1200},
}));


// Capa tem page break (única exceção)
children.push(pgBreak());

// ── 1. APRESENTAÇÃO ───────────────────────────────────────────────
children.push(
  secHead("Apresentação","1"),
  para('Este relatório consolida as informações de ' + INFO.num_locais_extenso +
    ' (' + DADOS.length + ') locais de aplicação das provas objetivas do ' + INFO.titulo +
    ' — ' + INFO.edital + ', realizadas em ' + INFO.data + ', turno ' + INFO.turno +
    '. Os ' + DADOS.length + ' locais dispõem de relatório técnico individual, arquivado separadamente.'),
  para('Os quantitativos de candidatos inscritos, presentes e ausentes apresentados ' +
    'neste documento têm como fonte autoritativa a base de leitura do IBGP, totalizando ' +
    totalCargos + ' cargos distribuídos em ' + totalSalas + ' salas de prova.'),
  space(60)
);

// ── 2. RESUMO EXECUTIVO ───────────────────────────────────────────
children.push(secHead("Resumo Executivo","2"));

children.push(
  para('A aplicação foi conduzida simultaneamente em ' + DADOS.length +
    ' locais, totalizando ' + totalSalas + ' salas. Foram registrados ' +
    totInscr.toLocaleString('pt-BR') + ' candidatos inscritos, ' +
    totPres.toLocaleString('pt-BR') + ' presentes e ' +
    totAus.toLocaleString('pt-BR') + ' ausentes, com taxa global de comparecimento de ' +
    pctPres + '% e ausência de ' + pctAus + '%.'),
  para('Em todos os locais a aplicação transcorreu conforme o cronograma: ' +
    'abertura do portão às ' + HR.ab_portao + ', fechamento do portão às ' + HR.fch_portao +
    ' e início das provas às ' + HR.inicio + '.'),
  space(100)
);

// ── Indicadores Globais (número grande + label abaixo — igual à imagem) ──
children.push(
  new Paragraph({
    children:[new TextRun({text:"Indicadores Globais", bold:true, size:22, color:ESC})],
    border:{ left:{style:BorderStyle.THICK, size:16, color:VRM} },
    indent:{left:240}, spacing:{before:60, after:120},
  })
);

var iW = 2250; // 4 × 2250 = 9000

function indicCell(val, label, fill) {
  return new TableCell({
    children: [
      new Paragraph({
        children:[new TextRun({text:val, bold:true, size:52, color:BR})],
        alignment: AlignmentType.CENTER,
        spacing: {before:120, after:40},
      }),
      new Paragraph({
        children:[new TextRun({text:label, bold:false, size:14, color:BR})],
        alignment: AlignmentType.CENTER,
        spacing: {before:0, after:120},
      }),
    ],
    shading: {type:ShadingType.CLEAR, fill:fill},
    borders: {top:BNONE, bottom:BNONE, left:BNONE, right:BNONE},
    width: {size:iW, type:WidthType.DXA},
    margins: {top:60, bottom:60, left:60, right:60},
  });
}

children.push(new Table({
  rows: [
    new TableRow({ children:[
      indicCell(String(DADOS.length),                        "LOCAIS DE APLICAÇÃO",    IBLUE),
      indicCell(String(totalCargos),                         "CARGOS",                 ITEAL),
      indicCell(String(totalSalas),                          "SALAS UTILIZADAS",       INDIGO),
      indicCell(totInscr.toLocaleString('pt-BR'),            "CANDIDATOS INSCRITOS",   IGRN),
    ]}),
    new TableRow({ children:[
      indicCell(totPres.toLocaleString('pt-BR'),             "CANDIDATOS PRESENTES",   IGRN),
      indicCell(totAus.toLocaleString('pt-BR'),              "CANDIDATOS AUSENTES",    IRED),
      indicCell(pctPres + '%',                               "TAXA DE PRESENÇA",       ISLAT),
      indicCell(pctAus + '%',                                "TAXA DE AUSÊNCIA",       DKRED),
    ]}),
  ],
  width:{size:iW*4, type:WidthType.DXA}, columnWidths:[iW,iW,iW,iW],
}));
children.push(space(180));

// Top 5 inscritos por cargo
children.push(
  subHead("Top 5 Cargos com Maior Número de Inscritos"),
  makeTable(
    ["Cargo / Função","Inscritos","Presentes","Ausentes","% Aus."],
    top5inscr.map(function(e){ return [e.nome, e.inscr, e.pres, e.aus,
      e.inscr?(e.aus/e.inscr*100).toFixed(1)+'%':'0.0%']; }),
    [5400,900,900,900,900], {}
  ),
  space(140)
);

// Top 5 ausência por cargo
children.push(
  subHead("Top 5 Cargos com Maior Taxa de Ausência"),
  makeTable(
    ["Cargo / Função","Inscritos","Ausentes","% Ausência"],
    top5cargos.map(function(c){ return [c.cargo, c.inscr, c.aus, c.pct.toFixed(1)+'%']; }),
    [6100,900,900,1100],
    { highCol:3 }
  ),
  space(140)
);

// Procedimentos padronizados
children.push(
  subHead("Procedimentos Padronizados — Aplicados em Todos os Locais"),
  bullet("Reunião de alinhamento com a equipe: " + HR.reuniao + "."),
  bullet("Abertura do portão às " + HR.ab_portao + ", com conferência de documentos e lista de presença da equipe."),
  bullet("Abertura dos malotes de provas com testemunhas e registro fotográfico."),
  bullet("Distribuição dos pacotes de provas às " + HR.distrib + ", com capa virada para a carteira."),
  bullet("Fechamento do portão pontualmente às " + HR.fch_portao + ", com lavratura do Termo de Fechamento."),
  bullet("Início das provas às " + HR.inicio + ", conforme cronograma do edital."),
  bullet("Identificação dos candidatos mediante documento oficial com foto."),
  bullet("Proibição e recolhimento de dispositivos eletrônicos antes do início."),
  bullet("Aplicação de prova condicional com protocolo e registro em formulário."),
  bullet("Fechamento e lacre dos malotes com folhas de respostas, sob testemunho dos coordenadores locais."),
  bullet("Encaminhamento de formulários, atas e malotes à coordenação logística."),
  space(60)
);

// ── 3. LOCAIS DE APLICAÇÃO ────────────────────────────────────────
children.push(
  secHead("Locais de Aplicação","3"),
  para("A tabela abaixo lista os " + DADOS.length + " locais de aplicação das provas, com o número de salas utilizadas e a coordenação local responsável."),
  space(80)
);

var locRows = DADOS.map(function(e,i){
  return [String(i+1), e.nome, INFO.data, String(e.salas), e.coord_local];
});
locRows.push(["","TOTAL — " + DADOS.length + " LOCAIS", INFO.data, String(totalSalas), ""]);

children.push(
  makeTable(
    ["#","Local de Aplicação","Data","Salas","Coord. Local"],
    locRows,
    [400,4000,1000,600,3000],
    { totalRow:true }
  ),
  space(140)
);

// ── 4. QUADRO CONSOLIDADO POR CARGO ──────────────────────────────
children.push(
  secHead("Quadro Consolidado por Cargo","4"),
  para("Consolidação dos " + totalCargos + " cargos do certame. Linhas com taxa de ausência igual ou superior a 50% são destacadas em vermelho."),
  space(80)
);

var cargoList = Object.entries(cargoMap).sort();
var cargoRows = cargoList.map(function(kv,i){
  var d=kv[1], nm=kv[0].replace(/^01\s+/,'');
  var pct = d.inscr?(d.aus/d.inscr*100).toFixed(1)+'%':'0.0%';
  return [String(i+1), nm, d.inscr, d.pres, d.aus, d.inc, pct];
});
cargoRows.push(["","TOTAL GERAL",totInscr,totPres,totAus,totInc,pctAus+'%']);

children.push(
  makeTable(
    ["#","Cargo / Função","Inscritos","Presentes","Ausentes","Inclusão","% Aus."],
    cargoRows,
    [400,4300,900,900,900,700,900],
    { totalRow:true, highCol:6 }
  ),
  space(140)
);

// ── 5. RESULTADO POR LOCAL ────────────────────────────────────────
children.push(
  secHead("Resultado por Local","5"),
  para("Resultado consolidado por local de aplicação: salas utilizadas, quantitativos, taxa de ausência, ocorrências e cargo com maior taxa de ausência."),
  space(80)
);

var resRows = DADOS.map(function(e,i){
  var t=escolaTot(e);
  var pct = t.inscr?(t.aus/t.inscr*100).toFixed(1)+'%':'0.0%';
  var ocorr = (e.ocorrencias && !/nenhuma/i.test(e.ocorrencias)) ? 'Sim' : 'Não';
  return [String(i+1), e.nome, e.salas, t.inscr, t.pres, t.aus, pct, ocorr, topCargoAus(e)];
});
resRows.push(["","TOTAL GERAL",totalSalas,totInscr,totPres,totAus,pctAus+'%','','']);

children.push(
  makeTable(
    ["#","Local","Salas","Prev.","Pres.","Aus.","% Aus.","Ocorr.","Cargo c/ Maior Aus."],
    resRows,
    [300,2400,550,600,600,600,700,550,2700],
    { totalRow:true, highCol:6 }
  ),
  space(140)
);

// ── 6. DETALHAMENTO OPERACIONAL POR LOCAL ─────────────────────────
children.push(
  secHead("Detalhamento Operacional por Local","6"),
  para("Para cada local são apresentadas as informações operacionais e os quantitativos por cargo."),
  space(60)
);

DADOS.forEach(function(e,idx){
  var t = escolaTot(e);

  children.push(new Table({
    rows:[new TableRow({children:[new TableCell({
      children:[new Paragraph({
        children:[new TextRun({text:(idx+1)+'. '+e.nome, bold:true, size:20, color:BR})],
        alignment: AlignmentType.LEFT,
        spacing:{before:0,after:0},
      })],
      shading:{type:ShadingType.CLEAR, fill:ESC},
      borders:{top:BNONE,bottom:BNONE,left:BNONE,right:BNONE},
      width:{size:9000,type:WidthType.DXA},
      margins:{top:100,bottom:100,left:160,right:160},
    })]})],
    width:{size:9000,type:WidthType.DXA}, columnWidths:[9000],
    spacing:{before:220,after:0},
  }));

  children.push(kvTable([
    ["Data da aplicação",     INFO.data],
    ["Salas utilizadas",      e.salas + ' sala' + (e.salas!==1?'s':'')],
    ["Coordenador Pedagógico",INFO.coord_ped],
    ["Coordenador Logístico", INFO.coord_log],
    ["Coordenação Local",     e.coord_local],
    ["Turno",                 e.turno],
    ["Andamento da prova",    e.andamento],
    ["Toque de celular",      e.toque_celular],
    ["Ocorrências",           e.ocorrencias || "Nenhuma ocorrência registrada."],
  ], {w1:2700, w2:6300}));

  children.push(space(60));

  var eMap = {};
  ['manha','tarde'].forEach(function(turno){
    (e[turno]||[]).forEach(function(r){
      if (!eMap[r[1]]) eMap[r[1]]={inscr:0,pres:0,aus:0,inc:0};
      eMap[r[1]].inscr+=r[2]; eMap[r[1]].pres+=r[3]; eMap[r[1]].aus+=r[4]; eMap[r[1]].inc+=r[5];
    });
  });
  var eRows = Object.entries(eMap).sort().map(function(kv){
    var d=kv[1], nm=kv[0].replace(/^01\s+/,'');
    var pct = d.inscr?(d.aus/d.inscr*100).toFixed(1)+'%':'0.0%';
    return [nm, d.inscr, d.pres, d.aus, d.inc, pct];
  });
  eRows.push(["Subtotal",t.inscr,t.pres,t.aus,t.inc,
    t.inscr?(t.aus/t.inscr*100).toFixed(1)+'%':'0.0%']);

  children.push(
    makeTable(
      ["Cargo / Função","Previstos","Presentes","Ausentes","Inclusão","% Ausência"],
      eRows, [4500,900,900,900,900,900],
      { totalRow:true, highCol:5 }
    ),
    space(120)
  );
});

// ── 7. REGISTRO FOTOGRÁFICO ───────────────────────────────────────
children.push(
  secHead("Registro Fotográfico por Local","7"),
  para("Imagens coletadas pelos coordenadores locais durante a aplicação, documentando: Sinalização da Escola, Abertura do Malote, Andamento da Aplicação e Fechamento do Malote."),
  space(80)
);

DADOS.forEach(function(e,idx){
  children.push(
    new Paragraph({
      children:[new TextRun({text:(idx+1)+'. '+e.nome, bold:true, size:24, color:VRM})],
      border:{ left:{style:BorderStyle.THICK, size:16, color:VRM} },
      indent:{left:240}, spacing:{before:160,after:60},
    }),
    para("[ Fotos a serem inseridas a partir do formulário individual de cada local ]",{italic:true,after:30}),
    para("Categorias: Sinalização da Escola  |  Abertura do Malote  |  Andamento da Aplicação  |  Fechamento do Malote",{after:80})
  );
});

// ── GERAR ─────────────────────────────────────────────────────────
var doc = new Document({
  numbering: { config:[{
    reference:"blist",
    levels:[{ level:0, format:LevelFormat.BULLET, text:"•",
      alignment:AlignmentType.LEFT,
      style:{ paragraph:{ indent:{left:720,hanging:260} } } }],
  }]},
  sections:[{
    properties:{ page:{ margin:{top:1008,right:1008,bottom:1008,left:1008} } },
    children: children,
  }],
  styles:{
    default:{ document:{ run:{ font:"Times New Roman", size:24, color:"333333" } } },
  },
});

Packer.toBuffer(doc).then(function(buf){
  fs.writeFileSync(process.env.DOCX_OUTPUT, buf);
  console.log('Word gerado:', process.env.DOCX_OUTPUT);
}).catch(function(err){
  console.error('Erro:', err.message);
  process.exit(1);
});












"""


# ─────────────────────────────────────────────────────────────────
# LEITURA DA PLANILHA EXCEL
# ─────────────────────────────────────────────────────────────────

def load_from_excel(xlsx_path):
    """
    Lê planilha_ibgp.xlsx e retorna (concurso_info, escolas).

    Abas esperadas:
      CONCURSO   — campos chave/valor (título, edital, datas, coords, horários)
      LOCAIS     — uma linha por escola (nome, coord_local, salas, turno, andamento, ...)
      RESULTADOS — uma linha por cargo (escola, sala, turno, cargo, inscr, pres, aus, elim)
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ── Aba CONCURSO ─────────────────────────────────────────────
    campo_map = {
        "título do concurso":        "titulo",
        "edital":                    "edital",
        "data da prova":             "data",
        "coordenador pedagógico":    "coord_ped",
        "coordenador logístico":     "coord_log",
        "hr reunião — início":       "hr_reuniao_ini",
        "hr reunião — fim":          "hr_reuniao_fim",
        "hr abertura do portão":     "hr_ab_portao",
        "hr distribuição de provas": "hr_distrib",
        "hr fechamento do portão":   "hr_fch_portao",
        "hr início da prova":        "hr_inicio_prova",
    }
    info = {
        "gerado_em": date.today().strftime("%d/%m/%Y"),
        "num_locais_extenso": "",
    }
    if "CONCURSO" in wb.sheetnames:
        ws = wb["CONCURSO"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue
            key = str(row[0]).strip().lower()
            v   = str(row[1]).strip() if row[1] is not None else ""
            if key in campo_map:
                info[campo_map[key]] = v
    # Preenche turno provisório (será calculado dinamicamente no JS)
    info.setdefault("turno", "")
    info.setdefault("titulo",         CONCURSO_INFO["titulo"])
    info.setdefault("edital",         CONCURSO_INFO["edital"])
    info.setdefault("data",           CONCURSO_INFO["data"])
    info.setdefault("coord_ped",      CONCURSO_INFO["coord_ped"])
    info.setdefault("coord_log",      CONCURSO_INFO["coord_log"])
    info.setdefault("hr_reuniao_ini", CONCURSO_INFO["hr_reuniao_ini"])
    info.setdefault("hr_reuniao_fim", CONCURSO_INFO["hr_reuniao_fim"])
    info.setdefault("hr_ab_portao",   CONCURSO_INFO["hr_ab_portao"])
    info.setdefault("hr_distrib",     CONCURSO_INFO["hr_distrib"])
    info.setdefault("hr_fch_portao",  CONCURSO_INFO["hr_fch_portao"])
    info.setdefault("hr_inicio_prova",CONCURSO_INFO["hr_inicio_prova"])

    # ── Aba LOCAIS ────────────────────────────────────────────────
    locais_meta = {}   # nome_lower → dict
    if "LOCAIS" in wb.sheetnames:
        ws = wb["LOCAIS"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            nome = str(row[0]).strip()
            if nome.startswith("⚠") or nome.startswith("#"): continue
            locais_meta[nome.lower()] = {
                "nome":          nome,
                "coord_local":   str(row[1]).strip() if row[1] else "",
                "salas":         int(row[2]) if isinstance(row[2], (int, float)) else 0,
                "turno":         str(row[3]).strip() if row[3] else "AMBOS",
                "andamento":     str(row[4]).strip() if row[4] else "",
                "toque_celular": str(row[5]).strip() if row[5] else "Não consta",
                "ocorrencias":   str(row[6]).strip() if row[6] else "",
                "manha":         [],
                "tarde":         [],
            }

    # ── Aba RESULTADOS ────────────────────────────────────────────
    if "RESULTADOS" in wb.sheetnames:
        ws = wb["RESULTADOS"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[3]: continue
            escola  = str(row[0]).strip()
            sala    = str(row[1]).strip() if row[1] else ""
            turno_r = str(row[2]).strip().lower() if row[2] else ""
            cargo   = str(row[3]).strip()
            nums    = [int(v) if isinstance(v, (int, float)) else 0 for v in row[4:8]]
            inscr, pres, aus, elim = nums[0], nums[1], nums[2], nums[3]

            key = escola.lower()
            if key not in locais_meta:
                # Escola mencionada em RESULTADOS mas não em LOCAIS — cria entrada básica
                locais_meta[key] = {
                    "nome": escola, "coord_local": "", "salas": 0, "turno": "AMBOS",
                    "andamento": "", "toque_celular": "Não consta", "ocorrencias": "",
                    "manha": [], "tarde": [],
                }
            turno_key = "tarde" if "tarde" in turno_r else "manha"
            locais_meta[key][turno_key].append((sala, cargo, inscr, pres, aus, elim))

    escolas = list(locais_meta.values())

    # Extenso do número de locais
    extensos = ["zero","um","dois","três","quatro","cinco","seis","sete","oito","nove","dez"]
    n = len(escolas)
    info["num_locais_extenso"] = extensos[n] if n < len(extensos) else str(n)

    return info, escolas


# ─────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json, subprocess

    # Detecta se foi passada uma planilha Excel
    xlsx_arg = None
    out_arg  = None
    for arg in sys.argv[1:]:
        if arg.endswith(".xlsx") or arg.endswith(".xls"):
            xlsx_arg = arg
        elif os.path.isdir(arg) or arg.endswith("/") or arg.endswith("\\"):
            out_arg = arg

    if xlsx_arg:
        print(f"Lendo dados de: {xlsx_arg}")
        info, escolas = load_from_excel(xlsx_arg)
        OUT_DIR = out_arg or os.path.dirname(os.path.abspath(xlsx_arg))
    else:
        print("Usando dados de exemplo (passe a planilha como argumento para dados reais)")
        info, escolas = CONCURSO_INFO, ESCOLAS
        OUT_DIR = out_arg or os.path.dirname(os.path.abspath(__file__))

    os.makedirs(OUT_DIR, exist_ok=True)
    PDF_OUT  = os.path.join(OUT_DIR, "relatorio_geral_aplicacao.pdf")
    DOCX_OUT = os.path.join(OUT_DIR, "relatorio_geral_aplicacao.docx")
    JS_FILE  = os.path.join(OUT_DIR, "relatorio_geral_gen.js")

    # Sempre (re)escreve o JS para não perder a versão correta
    with open(JS_FILE, "w", encoding="utf-8") as _f:
        _f.write(_DOCX_SCRIPT)

    # PDF
    gerar_pdf(PDF_OUT, info, escolas)

    # Word
    env = os.environ.copy()
    env["RELATORIO_INFO"]  = json.dumps(info,    ensure_ascii=False)
    env["RELATORIO_DADOS"] = json.dumps(escolas, ensure_ascii=False)
    env["DOCX_OUTPUT"]     = DOCX_OUT
    env["LOGO_PATH"]       = LOGO_PATH

    result = subprocess.run(["node", JS_FILE], env=env,
                            capture_output=True, text=True, cwd=OUT_DIR)
    if result.stdout: print(result.stdout.strip())
    if result.returncode != 0:
        print("Erro Word:", result.stderr[:500])

    print("\nArquivos gerados:")
    print(f"  PDF:  {PDF_OUT}")
    print(f"  Word: {DOCX_OUT}")
